import pandas as pd
import numpy as np
from datetime import datetime, time
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

def analyze_attendance_data(csv_file, output_file):
    """
    7월 근태 데이터를 분석하여 출근이 7:50 이후이고 퇴근이 17:10 이전인 셀을 하이라이트
    (8:00 정각과 17:00 정각은 제외)
    
    Args:
        csv_file (str): 입력 CSV 파일 경로
        output_file (str): 출력 Excel 파일 경로
    """
    
    # CSV 파일 읽기 (인코딩 문제 해결)
    try:
        df = pd.read_csv(csv_file, encoding='utf-8')
    except UnicodeDecodeError:
        try:
            df = pd.read_csv(csv_file, encoding='cp949')
        except UnicodeDecodeError:
            df = pd.read_csv(csv_file, encoding='euc-kr')
    
    print("데이터 로드 완료")
    print(f"데이터 형태: {df.shape}")
    print(f"컬럼명: {list(df.columns)}")
    
    # 시간 데이터를 파싱하는 함수
    def parse_time(time_str):
        """시간 문자열을 파싱하여 time 객체로 변환"""
        if pd.isna(time_str) or time_str == '':
            return None
        
        try:
            # 다양한 시간 형식 처리
            if isinstance(time_str, str):
                time_str = time_str.strip()
                if ':' in time_str:
                    return datetime.strptime(time_str, '%H:%M').time()
                else:
                    return None
            return None
        except:
            return None
    
    # 출근/퇴근 시간 기준 (8:00 정각과 17:00 정각은 제외)
    arrival_threshold = time(7, 50)  # 7:50
    departure_threshold = time(17, 10)  # 17:10
    
    # 새로운 워크북 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "근태분석결과"
    
    # 헤더 추가
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
    
    # 데이터 추가
    for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # 출근/퇴근 열 매핑 (F, H, J, L~)
    arrival_columns = []
    departure_columns = []
    
    for i, col in enumerate(df.columns):
        if i >= 5:  # F열부터 시작 (인덱스 5)
            if (i - 5) % 2 == 0:  # 짝수 인덱스: 출근 열
                arrival_columns.append(i)
            else:  # 홀수 인덱스: 퇴근 열
                departure_columns.append(i)
    
    print(f"출근 열 인덱스: {arrival_columns}")
    print(f"퇴근 열 인덱스: {departure_columns}")
    
    # 스타일 정의
    red_font = Font(color="FF0000", bold=True)  # 빨간색 글씨, 굵게
    pink_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # 분홍 배경
    
    # 각 행에 대해 시간 조건 확인 및 스타일 적용
    highlighted_cells = 0
    
    for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):  # Excel은 1부터 시작, 헤더 제외
        for col_idx in arrival_columns:
            if col_idx < len(row_data):
                arrival_time = parse_time(row_data.iloc[col_idx])
                departure_time = None
                
                # 해당하는 퇴근 시간 찾기
                if col_idx + 1 < len(row_data):
                    departure_time = parse_time(row_data.iloc[col_idx + 1])
                
                # 조건 확인: 
                # 1. 출근이 7:50 이후이고 8:00 정각이 아님
                # 2. 퇴근이 17:10 이전이고 17:00 정각이 아님
                if (arrival_time and 
                    arrival_time > arrival_threshold and 
                    arrival_time != time(8, 0) and  # 8:00 정각 제외
                    departure_time and 
                    departure_time < departure_threshold and
                    departure_time != time(17, 0)):  # 17:00 정각 제외
                    
                    # 출근 시간 셀 스타일 적용
                    cell = ws.cell(row=row_idx, column=col_idx + 1)
                    cell.font = red_font
                    cell.fill = pink_fill
                    
                    # 퇴근 시간 셀도 스타일 적용
                    if departure_time:
                        dep_cell = ws.cell(row=row_idx, column=col_idx + 2)
                        dep_cell.font = red_font
                        dep_cell.fill = pink_fill
                    
                    highlighted_cells += 1
                    print(f"행 {row_idx}, 열 {get_column_letter(col_idx + 1)}: 출근 {arrival_time}, 퇴근 {departure_time} - 하이라이트")
    
    print(f"\n총 {highlighted_cells}개의 셀이 하이라이트되었습니다.")
    
    # 열 너비 자동 조정
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 워크북 저장
    try:
        wb.save(output_file)
        print(f"결과가 {output_file}에 저장되었습니다.")
    except Exception as e:
        print(f"파일 저장 중 오류: {e}")
        # 백업 파일명으로 저장 시도
        backup_file = output_file.replace('.xlsx', '_backup.xlsx')
        wb.save(backup_file)
        print(f"백업 파일로 저장: {backup_file}")
    
    return highlighted_cells

def create_summary_report(csv_file):
    """근태 데이터 요약 리포트 생성"""
    
    # CSV 파일 읽기
    try:
        df = pd.read_csv(csv_file, encoding='utf-8')
    except UnicodeDecodeError:
        try:
            df = pd.read_csv(csv_file, encoding='cp949')
        except UnicodeDecodeError:
            df = pd.read_csv(csv_file, encoding='euc-kr')
    
    print("\n=== 7월 근태 데이터 요약 ===")
    print(f"총 직원 수: {len(df)}")
    print(f"데이터 컬럼 수: {len(df.columns)}")
    
    # 출근/퇴근 시간 통계
    arrival_times = []
    departure_times = []
    
    for i, col in enumerate(df.columns):
        if i >= 5:  # F열부터 시작
            if (i - 5) % 2 == 0:  # 출근 열
                for time_val in df.iloc[:, i]:
                    if pd.notna(time_val) and str(time_val).strip() != '':
                        try:
                            if ':' in str(time_val):
                                time_obj = datetime.strptime(str(time_val).strip(), '%H:%M').time()
                                arrival_times.append(time_obj)
                        except:
                            pass
            else:  # 퇴근 열
                for time_val in df.iloc[:, i]:
                    if pd.notna(time_val) and str(time_val).strip() != '':
                        try:
                            if ':' in str(time_val):
                                time_obj = datetime.strptime(str(time_val).strip(), '%H:%M').time()
                                departure_times.append(time_obj)
                        except:
                            pass
    
    if arrival_times:
        print(f"출근 시간 범위: {min(arrival_times)} ~ {max(arrival_times)}")
        print(f"평균 출근 시간: {sum([t.hour * 60 + t.minute for t in arrival_times]) / len(arrival_times) / 60:.2f}시")
    
    if departure_times:
        print(f"퇴근 시간 범위: {min(departure_times)} ~ {max(departure_times)}")
        print(f"평균 퇴근 시간: {sum([t.hour * 60 + t.minute for t in departure_times]) / len(departure_times) / 60:.2f}시")

def create_simple_excel(csv_file, output_file):
    """간단한 Excel 파일 생성 (문제 해결용)"""
    try:
        # CSV 파일 읽기
        df = pd.read_csv(csv_file, encoding='cp949')
        
        # Excel 파일로 저장 (기본 pandas 방식)
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='근태데이터')
            
            # 워크북 가져오기
            workbook = writer.book
            worksheet = writer.sheets['근태데이터']
            
            # 스타일 적용
            red_font = Font(color="FF0000", bold=True)
            pink_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            
            # 출근/퇴근 열 매핑
            arrival_columns = []
            for i in range(5, len(df.columns), 2):  # F, H, J, L...
                arrival_columns.append(i)
            
            # 조건에 맞는 셀 하이라이트
            highlighted_count = 0
            
            for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
                for col_idx in arrival_columns:
                    if col_idx < len(row_data):
                        arrival_val = row_data.iloc[col_idx]
                        departure_val = row_data.iloc[col_idx + 1] if col_idx + 1 < len(row_data) else None
                        
                        # 시간 파싱 및 조건 확인
                        try:
                            if pd.notna(arrival_val) and ':' in str(arrival_val):
                                arrival_time = datetime.strptime(str(arrival_val).strip(), '%H:%M').time()
                                
                                # 출근 시간이 7:50 이후이고 8:00 정각이 아닌 경우
                                if (arrival_time > time(7, 50) and arrival_time != time(8, 0) and 
                                    pd.notna(departure_val) and ':' in str(departure_val)):
                                    
                                    departure_time = datetime.strptime(str(departure_val).strip(), '%H:%M').time()
                                    
                                    # 퇴근 시간이 17:10 이전이고 17:00 정각이 아닌 경우
                                    if departure_time < time(17, 10) and departure_time != time(17, 0):
                                        
                                        # 출근 시간 셀
                                        cell = worksheet.cell(row=row_idx, column=col_idx + 1)
                                        cell.font = red_font
                                        cell.fill = pink_fill
                                        
                                        # 퇴근 시간 셀
                                        dep_cell = worksheet.cell(row=row_idx, column=col_idx + 2)
                                        dep_cell.font = red_font
                                        dep_cell.fill = pink_fill
                                        
                                        highlighted_count += 1
                                        print(f"행 {row_idx}, 열 {get_column_letter(col_idx + 1)}: 출근 {arrival_time}, 퇴근 {departure_time} - 하이라이트")
                                        
                        except Exception as e:
                            print(f"시간 파싱 오류 (행 {row_idx}, 열 {col_idx}): {e}")
                            pass
            
            print(f"간단한 Excel 파일 생성 완료: {highlighted_count}개 셀 하이라이트")
            
        return highlighted_count
        
    except Exception as e:
        print(f"간단한 Excel 생성 중 오류: {e}")
        return 0

if __name__ == "__main__":
    # 파일 경로 설정
    csv_file = "7월 근태 - 시트1.csv"
    output_file = "7월_근태_분석결과_개선.xlsx"
    
    try:
        # 요약 리포트 생성
        create_summary_report(csv_file)
        
        print("\n=== Excel 파일 생성 시도 ===")
        
        # 먼저 간단한 방식으로 시도
        print("1. 간단한 Excel 생성 방식 시도...")
        simple_count = create_simple_excel(csv_file, output_file)
        
        if simple_count > 0:
            print(f"✅ 간단한 방식으로 성공! 하이라이트된 셀: {simple_count}개")
            print(f"💾 결과 파일: {output_file}")
        else:
            print("❌ 간단한 방식 실패, 상세 분석 방식 시도...")
            # 상세 분석 방식 시도
            detailed_count = analyze_attendance_data(csv_file, output_file.replace('.xlsx', '_상세.xlsx'))
            print(f"✅ 상세 분석 방식 완료! 하이라이트된 셀: {detailed_count}개")
        
    except Exception as e:
        print(f"❌ 오류 발생: {str(e)}")
        import traceback
        traceback.print_exc()
